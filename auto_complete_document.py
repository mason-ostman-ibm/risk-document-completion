#!/usr/bin/env python3
"""
Auto Complete Document Tool
Automatically detects Q&A columns in Excel sheets and fills in answers using RAG
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import openpyxl.utils
from io import BytesIO
import os
import re
from ibm_watsonx_ai import Credentials
from ibm_watsonx_ai.foundation_models import ModelInference
from dotenv import load_dotenv
from astrapy import DataAPIClient
from sentence_transformers import SentenceTransformer

load_dotenv()

# Initialize AstraDB for RAG
ASTRA_DB_API_ENDPOINT = os.getenv("ASTRA_DB_API_ENDPOINT")
ASTRA_DB_APPLICATION_TOKEN = os.getenv("ASTRA_DB_APPLICATION_TOKEN")

astra_client = DataAPIClient(ASTRA_DB_APPLICATION_TOKEN)
astra_database = astra_client.get_database(ASTRA_DB_API_ENDPOINT)
collection = astra_database.get_collection("qa_collection")

# Initialize embedding model for RAG
embedding_model = SentenceTransformer('ibm-granite/granite-embedding-30m-english')

def initialize_model():
    """Initialize the LLM model for both column detection and question answering"""
    credentials = Credentials(
        url=os.getenv("MODEL_URL"),
        username="mason.ostman@ibm.com",
        api_key=os.getenv("API_KEY")
    )

    project_id = os.getenv("PROJECT_ID")
    space_id = os.getenv("SPACE_ID")
    model_id = os.getenv("MODEL")

    parameters = {
        "temperature": 0,
        "top_p": 1
    }

    model = ModelInference(
        model_id=model_id,
        params=parameters,
        credentials=credentials,
        project_id=project_id,
        space_id=space_id
    )

    return model

def detect_qa_columns_in_sheet(df, model):
    """
    Use LLM to detect which columns contain questions and answers

    Args:
        df: pandas DataFrame of the sheet
        model: Initialized LLM model

    Returns:
        tuple: (question_column_name, answer_column_name) or (None, None) if detection fails
    """
    sample_data = df.head(5).to_string()

    prompt = f"""Given this spreadsheet data, identify which column contains questions and which contains answers.

{sample_data}

Respond in this exact format:
Question column: [column name]
Answer column: [column name]"""

    messages = [
        {
            "role": "user",
            "content": prompt
        }
    ]

    try:
        response = model.chat(messages=messages)
        answer = response["choices"][0]["message"]["content"]

        # Parse the response
        question_match = re.search(r'Question column:\s*(.+)', answer)
        answer_match = re.search(r'Answer column:\s*(.+)', answer)

        if question_match and answer_match:
            question_col = question_match.group(1).strip()
            answer_col = answer_match.group(1).strip()
            return question_col, answer_col

        return None, None
    except Exception as e:
        print(f"Error detecting columns: {e}")
        return None, None

def get_relevant_context(question, top_k=5, similarity_threshold=0.5):
    """
    Search for relevant Q&A pairs using RAG

    Args:
        question: The user's question
        top_k: Number of top matches to return
        similarity_threshold: Minimum similarity score to include

    Returns:
        String formatted context ready to add to LLM prompt
    """
    query_embedding = embedding_model.encode(question).tolist()

    results = collection.find(
        sort={"$vector": query_embedding},
        limit=top_k,
        projection={"question": 1, "answer": 1, "category": 1, "source_file": 1},
        include_similarity=True
    )

    context = ""
    relevant_count = 0

    for j, result in enumerate(results, 1):
        similarity = result.get('$similarity', 0)

        if similarity < similarity_threshold:
            continue

        q = result.get('question', 'N/A')
        a = result.get('answer', 'N/A')

        answer_lower = str(a).lower().strip()
        if answer_lower in ['unanswered', 'nan', 'none', '', 'n/a']:
            continue

        relevant_count += 1
        context += f"Example {relevant_count}:\n"
        context += f"Q: {q}\n"
        context += f"A: {a}\n\n"

    if not context:
        context = "No relevant examples found. Use your general knowledge about IBM and business practices."

    return context.strip()

def ask_llm(question, model):
    """
    Ask the LLM a question with RAG context

    Args:
        question: The question to answer
        model: Initialized LLM model

    Returns:
        Generated answer
    """
    try:
        relevant_context = get_relevant_context(question, top_k=5, similarity_threshold=0.5)
    except Exception as e:
        relevant_context = "No context available due to retrieval error."

    temp_message = [
        {
            "role": "system",
            "content": """You are a professional document completion assistant for IBM. Your role is to fill out governance, compliance, and business documents with accurate, concise information.

INSTRUCTIONS:
1. You are filling out forms and documents on behalf of IBM
2. Answer questions directly and professionally, as if completing an official form
3. Use the provided context examples as reference for style, format, and content
4. Match the tone and detail level of the context examples
5. Be concise - forms require brief, direct answers, not explanations
6. If context is provided, adapt it to answer the specific question
7. Do NOT mention that you're using context or reference materials
8. Do NOT include meta-commentary like "based on the context" or "according to the information provided"
9. NEVER respond with just "unanswered" or leave the answer blank
10. If you don't have relevant information, write "Information not available" rather than making up content or writing "unanswered"

FORMATTING:
- For yes/no questions: Answer "Yes" or "No" followed by brief details if needed
- For descriptive questions: Provide 1-3 sentences maximum unless more detail is clearly needed
- For lists: Use bullet points or numbered lists as appropriate
- Maintain professional business language throughout

Remember: You ARE the person filling out this document. Write answers directly as they should appear in the form."""
        },
        {
        "role": "user",
        "content": f"""Using the following reference examples, answer the question below.

REFERENCE EXAMPLES:
{relevant_context}

QUESTION TO ANSWER:
{question}

Provide your answer:"""
        }
    ]

    try:
        generated_response = model.chat(messages=temp_message)
        answer = generated_response["choices"][0]["message"]["content"]

        if not answer or answer.strip() == "":
            return "Error: Empty response from model"

        return answer

    except Exception as e:
        return f"Error: {str(e)}"

def process_document(file_path, output_path=None):
    """
    Main function to process Excel document:
    1. Iterate through each sheet
    2. Detect Q&A columns using LLM
    3. Answer unanswered questions using RAG
    4. Save completed document

    Args:
        file_path: Path to input Excel file
        output_path: Path to output Excel file (optional)

    Returns:
        Path to output file
    """
    print(f"\n{'='*60}")
    print("AUTO COMPLETE DOCUMENT TOOL")
    print(f"{'='*60}\n")

    # Initialize model
    print("Initializing LLM model...")
    model = initialize_model()
    print("Model initialized!\n")

    # Load workbook
    print(f"Loading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=False)
    print(f"Found {len(wb.sheetnames)} sheets\n")

    total_questions_answered = 0
    sheets_processed = 0

    # Iterate through each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*60}")
        print(f"Processing sheet: {sheet_name}")
        print(f"{'='*60}")

        ws = wb[sheet_name]

        # Read sheet as DataFrame
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        if df.empty:
            print(f"  ⊘ Skipping empty sheet")
            continue

        # Detect Q&A columns using LLM
        print(f"  Detecting Q&A columns...")
        question_col_name, answer_col_name = detect_qa_columns_in_sheet(df, model)

        if not question_col_name or not answer_col_name:
            print(f"  ⊘ Could not detect Q&A columns, skipping sheet")
            continue

        print(f"  ✓ Question column: {question_col_name}")
        print(f"  ✓ Answer column: {answer_col_name}")

        # Get column indices (1-based for openpyxl)
        try:
            question_col_idx = df.columns.get_loc(question_col_name) + 1
            answer_col_idx = df.columns.get_loc(answer_col_name) + 1
        except KeyError:
            print(f"  ⊘ Column names not found in DataFrame, skipping sheet")
            continue

        # Set column widths
        ws.column_dimensions[openpyxl.utils.get_column_letter(question_col_idx)].width = 60
        ws.column_dimensions[openpyxl.utils.get_column_letter(answer_col_idx)].width = 80

        # Iterate through rows and answer questions
        questions_in_sheet = 0
        for idx, row in df.iterrows():
            excel_row = idx + 2  # +2 because: +1 for header, +1 for 0-based to 1-based

            question = row.get(question_col_name)
            answer = row.get(answer_col_name)

            # Skip if no question
            if pd.isna(question) or str(question).strip() == "":
                continue

            question_str = str(question).strip()
            answer_str = str(answer).strip() if pd.notna(answer) else ""

            # Check if answer is missing or "unanswered"
            if not answer_str or answer_str.lower() in ['nan', 'unanswered', '']:
                print(f"  Answering row {excel_row}: {question_str[:60]}...")

                # Get answer from LLM with RAG
                generated_answer = ask_llm(question_str, model)

                # Unmerge cells if needed before writing
                answer_cell = ws.cell(row=excel_row, column=answer_col_idx)
                if isinstance(answer_cell, openpyxl.cell.cell.MergedCell):
                    # Find and unmerge the range containing this cell
                    for merged_range in list(ws.merged_cells.ranges):
                        if answer_cell.coordinate in merged_range:
                            ws.unmerge_cells(str(merged_range))
                            break
                    # Get the cell again after unmerging
                    answer_cell = ws.cell(row=excel_row, column=answer_col_idx)

                # Fill in the answer
                answer_cell.value = generated_answer
                answer_cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Format question cell (handle merged cells)
                question_cell = ws.cell(row=excel_row, column=question_col_idx)
                if isinstance(question_cell, openpyxl.cell.cell.MergedCell):
                    # Find and unmerge the range containing this cell
                    for merged_range in list(ws.merged_cells.ranges):
                        if question_cell.coordinate in merged_range:
                            ws.unmerge_cells(str(merged_range))
                            break
                    # Get the cell again after unmerging
                    question_cell = ws.cell(row=excel_row, column=question_col_idx)

                question_cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Adjust row height
                estimated_lines = max(len(generated_answer) // 80, 1)
                min_height = max(15 * estimated_lines, 15)
                ws.row_dimensions[excel_row].height = min(min_height, 150)

                questions_in_sheet += 1
                total_questions_answered += 1

        sheets_processed += 1
        print(f"  ✓ Answered {questions_in_sheet} questions in this sheet")

    # Save workbook
    if output_path is None:
        from pathlib import Path
        input_path = Path(file_path)
        output_path = str(input_path.parent / f"{input_path.stem}_completed{input_path.suffix}")

    print(f"\n{'='*60}")
    print("PROCESSING SUMMARY")
    print(f"{'='*60}")
    print(f"Sheets processed: {sheets_processed}/{len(wb.sheetnames)}")
    print(f"Total questions answered: {total_questions_answered}")
    print(f"Output file: {output_path}")
    print(f"{'='*60}\n")

    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python auto_complete_document.py <input_file.xlsx> [output_file.xlsx]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    result = process_document(input_file, output_file)
    print(f"\n✓ Processing complete! File saved to: {result}")
