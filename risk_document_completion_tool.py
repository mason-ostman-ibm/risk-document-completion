#!/usr/bin/env python3
"""
Risk Document Completion Tool - MCP Server
Exposes RAG chatbot functionality as MCP tools for Claude and WatsonX Orchestrate

This MCP server provides tools for:
- Answering governance/compliance questions using RAG
- Processing Excel questionnaires with automated answers
- Searching the knowledge base for relevant Q&A pairs
- Getting statistics about the knowledge base
"""

from fastmcp import FastMCP
from pathlib import Path
from io import BytesIO
import base64
import os
import openpyxl
import openpyxl.utils
import pandas as pd
import warnings
import zipfile
import xml.etree.ElementTree as ET
import re
from typing import List, Tuple
from ibm_watsonx_ai import Credentials
from ibm_watsonx_ai.foundation_models import ModelInference
from dotenv import load_dotenv
from astrapy import DataAPIClient
from sentence_transformers import SentenceTransformer
from openpyxl.styles import Alignment
import sys

# Suppress benign openpyxl warnings about invalid specifications in Excel files
# These warnings are common with Excel files from different sources/versions
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# Global flag to control debug output (set based on transport mode)
_STDIO_MODE = False

def debug_print(*args, **kwargs):
    """
    Print debug/status messages to stderr in stdio mode, stdout in HTTP mode.
    In stdio mode, stdout must contain ONLY JSON-RPC messages.
    """
    if _STDIO_MODE:
        # In stdio mode, send all debug output to stderr
        print(*args, **kwargs, file=sys.stderr)
    else:
        # In HTTP mode, normal stdout is fine
        print(*args, **kwargs)

# Set up environment path - use local .env file in the same directory
SCRIPT_DIR = Path(__file__).parent
env_path = SCRIPT_DIR / '.env'

# Load environment variables from local .env file
load_dotenv(env_path)

# AstraDB Connection Setup
ASTRA_DB_API_ENDPOINT = os.getenv("ASTRA_DB_API_ENDPOINT")
ASTRA_DB_APPLICATION_TOKEN = os.getenv("ASTRA_DB_APPLICATION_TOKEN")

if not ASTRA_DB_API_ENDPOINT or not ASTRA_DB_APPLICATION_TOKEN:
    raise ValueError("Please set ASTRA_DB_API_ENDPOINT and ASTRA_DB_APPLICATION_TOKEN in .env file")

# Initialize AstraDB client
debug_print("Connecting to AstraDB...")
astra_client = DataAPIClient(ASTRA_DB_APPLICATION_TOKEN)
astra_database = astra_client.get_database(ASTRA_DB_API_ENDPOINT)
collection = astra_database.get_collection("qa_collection")
debug_print("Connected to AstraDB successfully!")

# Authorization
credentials = Credentials(
    url=os.getenv("MODEL_URL"),
    username="mason.ostman@ibm.com",
    api_key=os.getenv("API_KEY")
)
debug_print("api done")

project_id = os.getenv("PROJECT_ID")
space_id = os.getenv("SPACE_ID")
debug_print(project_id)

# Model initialization
model_id = os.getenv("MODEL")

embedding_model = SentenceTransformer('ibm-granite/granite-embedding-30m-english')
debug_print("Embedding model loaded")

parameters = {
    "frequency_penalty": 0,
    "max_tokens": 2000,
    "presence_penalty": 0,
    "temperature": 0,
    "top_p": 1
}

model = ModelInference(
    model_id=model_id,
    params=parameters,
    credentials=credentials,
    project_id=project_id,
    space_id=space_id
)


# ===== RAG Functions (from RAG_Chatbot.py) =====

def ask_llm(question, verbose=False):
    """
    Ask the LLM a question with RAG context

    Args:
        question: The question to answer
        verbose: Print debug information

    Returns:
        Generated answer
    """
    try:
        # Request more results since we filter out "unanswered" entries
        relevant_context = get_relevant_context(question, top_k=5, similarity_threshold=0.5, verbose=verbose)
    except Exception as e:
        if verbose:
            debug_print(f"Warning: Failed to get RAG context: {e}")
        relevant_context = "No context available due to retrieval error."

    # Improved system prompt for document filling
    temp_message = [
        {
            "role": "system",
            "content": """You are a professional document completion assistant for IBM. Your role is to fill out governance, compliance, and business documents with accurate, concise information.

INSTRUCTIONS:
1. You are filling out forms and documents on behalf of IBM
2. Answer questions directly and professionally, as if completing an official form
3. Use the provided context examples as reference for style, format, and content
4. Match the tone and detail level of the context examples
5. Be concise - forms require brief, direct answers, not explanations
6. If context is provided, adapt it to answer the specific question
7. Do NOT mention that you're using context or reference materials
8. Do NOT include meta-commentary like "based on the context" or "according to the information provided"
9. NEVER respond with just "unanswered" or leave the answer blank
10. If you don't have relevant information, write "Information not available" rather than making up content or writing "unanswered"

FORMATTING:
- For yes/no questions: Answer "Yes" or "No" followed by brief details if needed
- For descriptive questions: Provide 1-3 sentences maximum unless more detail is clearly needed
- For lists: Use bullet points or numbered lists as appropriate
- Maintain professional business language throughout

Remember: You ARE the person filling out this document. Write answers directly as they should appear in the form."""
        },
        {
        "role": "user",
        "content": f"""Using the following reference examples, answer the question below.

REFERENCE EXAMPLES:
{relevant_context}

QUESTION TO ANSWER:
{question}

Provide your answer:"""
        }
    ]

    try:
        generated_response = model.chat(messages=temp_message)
        answer = generated_response["choices"][0]["message"]["content"]

        if not answer or answer.strip() == "":
            return "Error: Empty response from model"

        return answer

    except KeyError as e:
        error_msg = f"Error: Unexpected response format from model - {str(e)}"
        if verbose:
            debug_print(error_msg)
        return error_msg
    except Exception as e:
        error_msg = f"Error: Model request failed - {type(e).__name__}: {str(e)[:100]}"
        if verbose:
            debug_print(error_msg)
        return error_msg


def get_relevant_context(question, top_k=3, similarity_threshold=0.5, verbose=False):
    """
    Search for relevant Q&A pairs to add to LLM prompt using AstraDB

    Args:
        question: The user's question
        top_k: Number of top matches to return (default: 3)
        similarity_threshold: Minimum similarity score to include (0-1, default: 0.5)
        verbose: Print debug information

    Returns:
        String formatted context ready to add to LLM prompt
    """
    # Embed the question
    query_embedding = embedding_model.encode(question).tolist()

    # Search AstraDB using vector similarity
    results = collection.find(
        sort={"$vector": query_embedding},
        limit=top_k,
        projection={"question": 1, "answer": 1, "category": 1, "source_file": 1},
        include_similarity=True
    )

    # Format results for LLM context
    context = ""
    relevant_count = 0

    for j, result in enumerate(results, 1):
        similarity = result.get('$similarity', 0)

        # Only include results above similarity threshold
        if similarity < similarity_threshold:
            if verbose:
                debug_print(f"  Skipping result {j}: similarity {similarity:.3f} below threshold {similarity_threshold}")
            continue

        q = result.get('question', 'N/A')
        a = result.get('answer', 'N/A')
        category = result.get('category', 'unknown')
        source = result.get('source_file', 'unknown')

        # Skip entries with "unanswered" or empty answers - these provide no useful context
        answer_lower = str(a).lower().strip()
        if answer_lower in ['unanswered', 'nan', 'none', '', 'n/a']:
            if verbose:
                debug_print(f"  Skipping result {j}: answer is '{a}' (not useful as context)")
            continue

        relevant_count += 1
        context += f"Example {relevant_count}:\n"
        context += f"Q: {q}\n"
        context += f"A: {a}\n"

        if verbose:
            debug_print(f"  ✓ Using result {j}: similarity {similarity:.3f}, category: {category}, source: {source}")

        context += "\n"

    if verbose:
        debug_print(f"  Found {relevant_count} relevant examples for context")

    if not context:
        context = "No relevant examples found. Use your general knowledge about IBM and business practices."

    return context.strip()


def should_skip_sheet(sheet_name):
    """
    Determine if a sheet should be skipped (instruction sheets, empty sheets, etc.)
    Uses same logic as document_upload_tool.py

    Args:
        sheet_name: Name of the worksheet

    Returns:
        True if sheet should be skipped, False otherwise
    """
    if not sheet_name:
        return True

    sheet_lower = sheet_name.lower()

    # Skip instruction/system sheets (same keywords as document_upload_tool.py)
    skip_keywords = ['instruction', 'dv_sheet', 'legend']

    return any(skip in sheet_lower for skip in skip_keywords)


def get_sheet_names_xml(file_path) -> List[str]:
    """
    Get sheet names from Excel file using XML parser
    Fallback method for files that openpyxl can't read properly

    Args:
        file_path: Path to Excel file

    Returns:
        List of sheet names
    """
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            wb_xml = zip_ref.read('xl/workbook.xml').decode('utf-8')
            sheet_info = re.findall(r'<sheet.*?name="([^"]+)"', wb_xml)
            return sheet_info
    except Exception:
        return []


def read_sheet_xml(file_path, sheet_name, question_column=1, answer_column=2) -> List[Tuple[int, str, bool]]:
    """
    Read Q&A data from Excel sheet using XML parser
    Fallback method for files that openpyxl/pandas can't read properly

    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet
        question_column: Column index for questions (1-based)
        answer_column: Column index for answers (1-based)

    Returns:
        List of tuples: [(row_num, question, has_answer), ...]
    """
    try:
        # Convert column numbers to letters
        def num_to_col_letter(n):
            result = ""
            while n > 0:
                n -= 1
                result = chr(65 + (n % 26)) + result
                n //= 26
            return result

        q_col_letter = num_to_col_letter(question_column)
        a_col_letter = num_to_col_letter(answer_column)

        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # Read shared strings
            try:
                ss_xml = zip_ref.read('xl/sharedStrings.xml')
                root = ET.fromstring(ss_xml)
                strings = []
                for si in root.findall('.//{http://purl.oclc.org/ooxml/spreadsheetml/main}si'):
                    t = si.find('.//{http://purl.oclc.org/ooxml/spreadsheetml/main}t')
                    if t is not None and t.text:
                        strings.append(t.text)
                    else:
                        strings.append('')
            except:
                strings = []

            # Find sheet file
            wb_xml = zip_ref.read('xl/workbook.xml').decode('utf-8')
            rels_xml = zip_ref.read('xl/_rels/workbook.xml.rels').decode('utf-8')

            sheet_info = re.findall(r'<sheet.*?name="([^"]+)".*?r:id="(rId\d+)"', wb_xml)
            sheet_rid = None
            for name, rid in sheet_info:
                if name == sheet_name:
                    sheet_rid = rid
                    break

            if not sheet_rid:
                return []

            # Get sheet file path
            rels = re.findall(rf'<Relationship.*?Id="{sheet_rid}".*?Target="([^"]+)"', rels_xml)
            if not rels:
                return []

            sheet_file = 'xl/' + rels[0]

            # Read sheet data
            sheet_xml = zip_ref.read(sheet_file)
            root = ET.fromstring(sheet_xml)

            qa_pairs = []
            rows = root.findall('.//{http://purl.oclc.org/ooxml/spreadsheetml/main}row')

            for row_elem in rows:
                row_num = int(row_elem.get('r', 0))
                if row_num == 0:
                    continue

                question = None
                answer = None

                # Get cells in this row
                for cell in row_elem.findall('.//{http://purl.oclc.org/ooxml/spreadsheetml/main}c'):
                    cell_ref = cell.get('r', '')
                    col = ''.join([c for c in cell_ref if c.isalpha()])

                    v_elem = cell.find('.//{http://purl.oclc.org/ooxml/spreadsheetml/main}v')
                    if v_elem is not None:
                        value = v_elem.text
                        cell_type = cell.get('t', 'n')

                        # Resolve shared string
                        if cell_type == 's' and strings:
                            try:
                                value = strings[int(value)]
                            except:
                                pass

                        if col == q_col_letter:
                            question = value
                        elif col == a_col_letter:
                            answer = value

                # Skip if no question or if it's a header row
                if question and str(question).strip() and str(question).lower() not in ['name', 'question']:
                    has_answer = bool(answer and str(answer).strip())
                    qa_pairs.append((row_num, str(question).strip(), has_answer))

            return qa_pairs

    except Exception as e:
        return []


def get_sheet_qa_data(file_path, sheet_name, question_column=1, answer_column=2):
    """
    Read Q&A data from a specific sheet using pandas (avoids phantom row issues)
    Falls back to XML parser for problematic files
    Similar to document_upload_tool.py approach

    Args:
        file_path: Path to Excel file
        sheet_name: Name of the sheet to read
        question_column: Column index for questions (1-based)
        answer_column: Column index for answers (1-based)

    Returns:
        List of tuples: [(row_num, question, has_answer), ...]
    """
    try:
        # Try pandas first (preferred method)
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # Check if we got valid data
        if df.empty or len(df) == 0:
            # Fallback to XML parser
            return read_sheet_xml(file_path, sheet_name, question_column, answer_column)

        # Convert to 0-based indexing for pandas
        q_col_idx = question_column - 1
        a_col_idx = answer_column - 1

        qa_pairs = []
        for idx, row in df.iterrows():
            # Get question and answer (use .get() to handle missing columns)
            question = row.get(q_col_idx)
            answer = row.get(a_col_idx)

            # Convert to string and check if valid
            question_str = str(question).strip() if pd.notna(question) else ""
            answer_str = str(answer).strip() if pd.notna(answer) else ""

            # Skip empty questions or pandas artifacts
            if not question_str or question_str == 'nan':
                continue

            # Check if answer exists (not empty, not 'nan', not 'unanswered')
            has_answer = bool(answer_str and answer_str not in ['nan', 'unanswered', ''])

            # Store: (excel_row_number, question, has_answer)
            # Add 1 to idx because pandas is 0-based but Excel rows are 1-based
            qa_pairs.append((idx + 1, question_str, has_answer))

        # If pandas returned no valid data, try XML parser
        if not qa_pairs:
            return read_sheet_xml(file_path, sheet_name, question_column, answer_column)

        return qa_pairs

    except Exception as e:
        # Fallback to XML parser on any error
        return read_sheet_xml(file_path, sheet_name, question_column, answer_column)


def process_excel(uploaded_file, question_column=1, answer_column=2, verbose=False, progress_callback=None):
    """
    Process Excel file and fill in answers using RAG

    Args:
        uploaded_file: Path to Excel file or file-like object
        question_column: Column index for questions (default: 1 = column A)
        answer_column: Column index for answers (default: 2 = column B)
        verbose: Print progress information
        progress_callback: Optional callback function(current, total, message) for progress updates

    Returns:
        BytesIO object with filled Excel file
    """
    def update_progress(current, total, message):
        """Helper to update progress"""
        if progress_callback:
            try:
                progress_callback(current, total, message)
            except Exception as e:
                if verbose:
                    debug_print(f"Warning: Progress callback failed: {e}")
        if verbose:
            debug_print(f"[{current}/{total}] {message}")

    # Validate inputs
    if question_column < 1:
        raise ValueError(f"Question column must be >= 1, got {question_column}")
    if answer_column < 1:
        raise ValueError(f"Answer column must be >= 1, got {answer_column}")
    if question_column == answer_column:
        raise ValueError(f"Question and answer columns must be different")

    # Load workbook with error handling
    # Note: We suppress openpyxl warnings but still catch actual errors
    try:
        # Use data_only=False to preserve formulas but get values
        # keep_vba=False to ignore VBA macros which can cause issues
        wb = openpyxl.load_workbook(uploaded_file, data_only=False, keep_vba=False)

        # Check if the file is corrupted (loads but has no accessible sheets)
        use_new_workbook = len(wb.sheetnames) == 0

        if use_new_workbook:
            if verbose:
                debug_print("⚠ File has corrupted sheet structure - will create new workbook")
                debug_print("  Original data will be read via XML parser")
            # Create a new workbook to write results
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
        else:
            if verbose:
                debug_print("✓ Excel file loaded successfully")
                debug_print("  (Note: Minor file format warnings were automatically handled)")

    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found: {uploaded_file}")
    except PermissionError:
        raise PermissionError(f"Permission denied accessing file: {uploaded_file}")
    except Exception as e:
        # Provide helpful error message for corrupted files
        error_msg = f"Failed to load Excel file. "
        error_msg += f"The file may be corrupted, password-protected, or in an unsupported format. "
        error_msg += f"Error: {str(e)}"
        raise Exception(error_msg)

    # First pass: count total questions using pandas (avoids phantom rows)
    update_progress(0, 100, "Analyzing workbook structure...")

    total_questions = 0
    sheet_data = {}  # Store sheet data to avoid re-reading

    # Get sheet names - use XML parser if openpyxl shows 0 sheets
    sheet_names = wb.sheetnames if wb.sheetnames else get_sheet_names_xml(uploaded_file)

    if not sheet_names:
        raise Exception("No sheets found in workbook. File may be corrupted or empty.")

    if verbose and not wb.sheetnames:
        debug_print("⚠ Using XML parser fallback for problematic file format")

    for sheet_name in sheet_names:
        if should_skip_sheet(sheet_name):
            if verbose:
                debug_print(f"⊘ Skipping sheet (keyword match): {sheet_name}")
            continue

        # Read Q&A data using pandas
        qa_pairs = get_sheet_qa_data(uploaded_file, sheet_name, question_column, answer_column)

        # Count unanswered questions
        unanswered = [(row_num, q) for row_num, q, has_answer in qa_pairs if not has_answer]

        if unanswered:
            sheet_data[sheet_name] = unanswered
            total_questions += len(unanswered)
            if verbose:
                debug_print(f"✓ Sheet '{sheet_name}': {len(unanswered)} unanswered questions")
        else:
            if verbose:
                debug_print(f"⊘ Skipping sheet (no unanswered questions): {sheet_name}")

    if total_questions == 0:
        update_progress(100, 100, "No questions found to answer")
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    update_progress(10, 100, f"Found {total_questions} questions across {len(sheet_data)} sheets")

    # Process all sheets (using pre-loaded data)
    total_answered = 0
    sheets_processed = 0
    current_progress = 10  # Start after analysis

    for sheet_name, unanswered_questions in sheet_data.items():
        # Try to get worksheet - create if doesn't exist (for problematic files)
        try:
            ws = wb[sheet_name]
        except KeyError:
            # Sheet exists in XML but not accessible via openpyxl
            # This happens with corrupted file formats
            # Create a new sheet and populate with ALL data from XML
            if verbose:
                debug_print(f"⚠ Creating new sheet '{sheet_name}' (original not accessible)")
            ws = wb.create_sheet(sheet_name)

            # Get ALL Q&A data from XML (including answered ones) to preserve structure
            all_qa_data = get_sheet_qa_data(uploaded_file, sheet_name, question_column, answer_column)

            # Populate the new sheet with questions (we'll fill answers below)
            for row_num, question, has_answer in all_qa_data:
                # Write question
                q_cell = ws.cell(row=row_num, column=question_column)
                q_cell.value = question
                q_cell.alignment = Alignment(wrap_text=True, vertical='top')

                # If there was already an answer in the original, preserve it
                if has_answer:
                    # Try to read the original answer from XML
                    # For now, we'll let the LLM fill it, or you could read from XML
                    pass

        # Set appropriate column widths for better readability
        # Only set if not already customized (width is default or very small)
        if ws.column_dimensions[openpyxl.utils.get_column_letter(question_column)].width <= 13:
            ws.column_dimensions[openpyxl.utils.get_column_letter(question_column)].width = 60
        if ws.column_dimensions[openpyxl.utils.get_column_letter(answer_column)].width <= 13:
            ws.column_dimensions[openpyxl.utils.get_column_letter(answer_column)].width = 80

        sheet_question_count = len(unanswered_questions)

        update_progress(current_progress, 100, f"Processing sheet: {sheet_name} ({sheet_question_count} questions)")

        sheet_answered = 0

        # Process each unanswered question
        for row_num, question_text in unanswered_questions:
            # Update progress for this specific question
            progress_pct = int(current_progress + ((sheet_answered / sheet_question_count) * (90 - current_progress) / len(sheet_data)))
            update_progress(
                progress_pct,
                100,
                f"Sheet '{sheet_name}': Question {sheet_answered + 1}/{sheet_question_count}"
            )

            try:
                # Get answer from LLM with RAG
                if verbose:
                    debug_print(f"  Processing Row {row_num}: {question_text[:80]}...")

                answer = ask_llm(question_text, verbose=verbose)

                # Validate answer
                if answer is None or answer == "":
                    answer = "Error: Empty response from LLM"

                # Fill in the answer in the Excel workbook
                answer_cell = ws.cell(row=row_num, column=answer_column)
                answer_cell.value = str(answer)
                answer_cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Also ensure the question cell has proper formatting
                question_cell = ws.cell(row=row_num, column=question_column)
                if not question_cell.alignment or not question_cell.alignment.wrap_text:
                    question_cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Auto-adjust row height based on content length
                # Estimate: ~50 chars per line, assume column width of 80
                estimated_lines = max(len(str(answer)) // 80, 1)
                min_height = max(15 * estimated_lines, 15)  # 15 points per line, min 15
                if ws.row_dimensions[row_num].height is None or ws.row_dimensions[row_num].height < min_height:
                    ws.row_dimensions[row_num].height = min(min_height, 150)  # Cap at 150 points

                sheet_answered += 1
                total_answered += 1

                if verbose:
                    debug_print(f"  ✓ Row {row_num}: {answer[:100]}...")

            except KeyboardInterrupt:
                # Allow user to cancel
                if verbose:
                    debug_print("\n\nProcessing cancelled by user")
                raise
            except Exception as e:
                # Log the error but continue processing
                error_msg = f"Error: {type(e).__name__} - {str(e)[:100]}"
                if verbose:
                    debug_print(f"  ✗ Row {row_num}: {error_msg}")

                answer_cell = ws.cell(row=row_num, column=answer_column)
                answer_cell.value = error_msg
                answer_cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Also ensure the question cell has proper formatting
                question_cell = ws.cell(row=row_num, column=question_column)
                if not question_cell.alignment or not question_cell.alignment.wrap_text:
                    question_cell.alignment = Alignment(wrap_text=True, vertical='top')

                sheet_answered += 1
                total_answered += 1

        sheets_processed += 1
        current_progress = int(10 + ((sheets_processed / len(sheet_data)) * 80))

        if verbose:
            debug_print(f"\n✓ Sheet '{sheet_name}': Answered {sheet_answered}/{sheet_question_count} questions")

    # Save workbook
    update_progress(90, 100, "Saving processed workbook...")

    # Print summary
    sheets_skipped = len(wb.sheetnames) - sheets_processed
    debug_print(f"\n{'='*60}")
    debug_print("PROCESSING SUMMARY")
    debug_print(f"{'='*60}")
    debug_print(f"Sheets processed: {sheets_processed}")
    debug_print(f"Sheets skipped: {sheets_skipped}")
    debug_print(f"Total questions answered: {total_answered}/{total_questions}")
    debug_print(f"{'='*60}\n")

    # Save to BytesIO with error handling
    try:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
    except PermissionError:
        raise PermissionError("Failed to save workbook: Permission denied")
    except Exception as e:
        raise Exception(f"Failed to save workbook: {str(e)}")

    update_progress(100, 100, f"Complete! Answered {total_answered}/{total_questions} questions")

    return output


# ===== MCP Server Setup =====

# Initialize MCP server
mcp = FastMCP("Risk Document Completion Tool")


@mcp.tool()
def answer_question(question: str) -> str:
    """
    Answer a governance/compliance/risk question using RAG.

    This tool uses Retrieval-Augmented Generation (RAG) to answer questions
    by searching the knowledge base for relevant Q&A pairs and using an LLM
    to generate accurate, context-aware answers.

    Args:
        question: The question to answer (e.g., "What is IBM's data privacy policy?")

    Returns:
        Generated answer based on knowledge base and LLM

    Example:
        answer_question("What is IBM's corporate headquarters address?")
    """
    try:
        answer = ask_llm(question, verbose=False)
        return answer
    except Exception as e:
        return f"Error: {str(e)}"


@mcp.tool()
def search_knowledge_base(query: str, top_k: int = 5, similarity_threshold: float = 0.5) -> str:
    """
    Search the vector database for relevant Q&A pairs.

    This tool performs semantic search to find the most relevant questions
    and answers from the knowledge base based on vector similarity.

    Args:
        query: Search query (e.g., "data privacy policies")
        top_k: Number of results to return (default: 5, max: 20)
        similarity_threshold: Minimum similarity score 0-1 (default: 0.5)

    Returns:
        Formatted string with matching Q&A pairs and similarity scores

    Example:
        search_knowledge_base("vendor risk assessment", top_k=3, similarity_threshold=0.6)
    """
    try:
        # Validate inputs
        if top_k < 1 or top_k > 20:
            return "Error: top_k must be between 1 and 20"

        if similarity_threshold < 0 or similarity_threshold > 1:
            return "Error: similarity_threshold must be between 0 and 1"

        context = get_relevant_context(
            query,
            top_k=top_k,
            similarity_threshold=similarity_threshold,
            verbose=True
        )
        return context
    except Exception as e:
        return f"Error: {str(e)}"


@mcp.tool()
def process_excel_questionnaire(
    file_path: str = "",
    file_base64: str = "",
    question_column: int = 2,
    answer_column: int = 3,
    output_path: str = ""
) -> str:
    """
    Process an Excel questionnaire file and fill in answers using RAG.

    This tool reads an Excel file with questions, uses the RAG system to
    generate answers for unanswered questions, and returns the result.

    TWO WAYS TO PROVIDE THE FILE:
    1. file_path: Provide absolute path to Excel file on the server
    2. file_base64: Provide base64-encoded Excel file content

    Args:
        file_path: Absolute path to Excel file (e.g., "/path/to/questionnaire.xlsx")
                   Leave empty if using file_base64
        file_base64: Base64-encoded Excel file content. Leave empty if using file_path
        question_column: Column number for questions (default: 2 = column B)
        answer_column: Column number for answers (default: 3 = column C)
        output_path: Optional output file path (default: auto-generated with "_processed" suffix)
                     Only used when file_path is provided

    Returns:
        If file_path provided: Status message with output file path and summary
        If file_base64 provided: JSON with base64-encoded result file and summary

    Examples:
        # Using file path (file already on server)
        process_excel_questionnaire(
            file_path="/Users/masonostman/Documents/Governance_Chatbot/my_rfp.xlsx",
            question_column=2,
            answer_column=3
        )

        # Using base64 (file sent by LLM/user)
        process_excel_questionnaire(
            file_base64="UEsDBBQABgAIAAAAIQ...",
            question_column=2,
            answer_column=3
        )
    """
    try:
        import json
        import tempfile

        # Validate inputs
        if not file_path and not file_base64:
            return "Error: Must provide either file_path or file_base64"

        if file_path and file_base64:
            return "Error: Provide either file_path OR file_base64, not both"

        # Validate column numbers
        if question_column < 1:
            return f"Error: question_column must be >= 1, got {question_column}"
        if answer_column < 1:
            return f"Error: answer_column must be >= 1, got {answer_column}"
        if question_column == answer_column:
            return f"Error: question_column and answer_column must be different"

        # Handle base64 input
        if file_base64:
            debug_print("Processing base64-encoded file...")
            try:
                # Decode base64
                file_data = base64.b64decode(file_base64)

                # Create temporary file
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                    tmp.write(file_data)
                    temp_file_path = tmp.name

                # Process the file
                debug_print(f"Question column: {question_column}, Answer column: {answer_column}")
                result_bytes = process_excel(
                    temp_file_path,
                    question_column=question_column,
                    answer_column=answer_column,
                    verbose=True
                )

                # Clean up temp file
                os.unlink(temp_file_path)

                # Encode result as base64
                result_base64 = base64.b64encode(result_bytes.getvalue()).decode('utf-8')
                file_size = len(result_bytes.getvalue())

                # Return JSON with base64 result
                response = {
                    "success": True,
                    "message": "File processed successfully",
                    "file_base64": result_base64,
                    "file_size_bytes": file_size,
                    "instructions": "Save the file_base64 content by decoding it and writing to a .xlsx file"
                }
                return json.dumps(response, indent=2)

            except base64.binascii.Error:
                return "Error: Invalid base64 encoding"
            except Exception as e:
                import traceback
                return f"Error processing base64 file: {str(e)}\n\n{traceback.format_exc()}"

        # Handle file path input (original behavior)
        else:
            # Validate file exists
            if not os.path.exists(file_path):
                return f"Error: File not found: {file_path}"

            # Validate file is Excel
            if not file_path.endswith(('.xlsx', '.xls')):
                return f"Error: File must be an Excel file (.xlsx or .xls): {file_path}"

            # Process the file
            debug_print(f"Processing file: {file_path}")
            debug_print(f"Question column: {question_column}, Answer column: {answer_column}")

            result_bytes = process_excel(
                file_path,
                question_column=question_column,
                answer_column=answer_column,
                verbose=True
            )

            # Determine output path
            if not output_path:
                input_path = Path(file_path)
                output_path = str(input_path.parent / f"{input_path.stem}_processed{input_path.suffix}")

            # Save the result
            with open(output_path, 'wb') as f:
                f.write(result_bytes.getvalue())

            file_size = len(result_bytes.getvalue())

            return f"Success! Processed file saved to: {output_path}\nFile size: {file_size:,} bytes"

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return f"Error: {str(e)}\n\nFull trace:\n{error_trace}"


@mcp.tool()
def get_database_stats() -> str:
    """
    Get statistics about the knowledge base (vector database).

    This tool provides insights into the knowledge base including:
    - Total number of documents
    - Documents by category (RFP, compliance, policy, etc.)
    - Source files in the database

    Returns:
        Statistics about the database contents

    Example:
        get_database_stats()
    """
    try:
        # Connect to AstraDB
        from astrapy import DataAPIClient
        from dotenv import load_dotenv

        # Load environment variables from local directory
        env_path = SCRIPT_DIR / '.env'
        load_dotenv(env_path)

        token = os.getenv('ASTRA_DB_APPLICATION_TOKEN')
        endpoint = os.getenv('ASTRA_DB_API_ENDPOINT')

        if not token or not endpoint:
            return "Error: AstraDB credentials not found in .env file"

        client = DataAPIClient(token)
        database = client.get_database(endpoint)
        collection = database.get_collection("qa_collection")

        # Get sample documents to analyze
        sample_docs = list(collection.find({}, limit=100))

        if not sample_docs:
            return "Database is empty"

        # Count by category
        categories = {}
        sources = set()

        for doc in sample_docs:
            cat = doc.get('category', 'unknown')
            categories[cat] = categories.get(cat, 0) + 1
            if 'source_file' in doc:
                sources.add(doc['source_file'])

        # Format response
        result = f"📊 Knowledge Base Statistics\n\n"
        result += f"Total documents sampled: {len(sample_docs)}\n"
        result += f"Unique source files: {len(sources)}\n\n"
        result += f"Documents by category:\n"
        for cat, count in sorted(categories.items()):
            result += f"  - {cat}: {count}\n"

        if sources:
            result += f"\nSource files:\n"
            for source in sorted(sources):
                result += f"  - {source}\n"

        return result

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return f"Error: {str(e)}\n\nFull trace:\n{error_trace}"


@mcp.tool()
def get_system_info() -> str:
    """
    Get information about the Risk Document Completion Tool system.

    Returns:
        System configuration and status information

    Example:
        get_system_info()
    """
    try:
        from dotenv import load_dotenv

        # Load environment variables from local directory
        env_path = SCRIPT_DIR / '.env'
        load_dotenv(env_path)

        result = "🔧 Risk Document Completion Tool - System Information\n\n"

        # Model configuration
        model_id = os.getenv("MODEL", "Not configured")
        result += f"LLM Model: {model_id}\n"
        result += f"Embedding Model: ibm-granite/granite-embedding-30m-english\n"

        # Database configuration
        db_endpoint = os.getenv("ASTRA_DB_API_ENDPOINT", "Not configured")
        if db_endpoint and len(db_endpoint) > 50:
            db_endpoint = db_endpoint[:50] + "..."
        result += f"Vector Database: AstraDB\n"
        result += f"Database Endpoint: {db_endpoint}\n"
        result += f"Collection: qa_collection\n"

        # Path configuration
        result += f"\nSystem Paths:\n"
        result += f"  MCP Server: {Path(__file__).resolve()}\n"
        result += f"  Environment File: {env_path}\n"

        result += f"\n✅ All RAG functions loaded successfully (inlined)\n"

        return result

    except Exception as e:
        return f"Error: {str(e)}"


# Run the server
if __name__ == "__main__":
    import sys

    # Set global flag for debug output routing based on transport mode
    if "--http" in sys.argv or "--sse" in sys.argv:
        # HTTP mode - can use stdout for debug messages
        _STDIO_MODE = False
    else:
        # stdio mode - must send debug to stderr only
        _STDIO_MODE = True

    debug_print("="*60)
    debug_print("Risk Document Completion Tool - MCP Server")
    debug_print("="*60)
    debug_print(f"Script Directory: {SCRIPT_DIR}")
    debug_print(f"Environment File: {SCRIPT_DIR / '.env'}")
    debug_print("="*60)

    # Check if HTTP mode is requested
    if "--http" in sys.argv or "--sse" in sys.argv:
        # HTTP/SSE mode for WatsonX Orchestrate
        port = 8080
        if "--port" in sys.argv:
            port_idx = sys.argv.index("--port") + 1
            if port_idx < len(sys.argv):
                port = int(sys.argv[port_idx])

        debug_print(f"\n🌐 Starting HTTP MCP server (SSE transport)")
        debug_print(f"   URL: http://0.0.0.0:{port}")
        debug_print(f"   SSE: http://localhost:{port}/sse")
        debug_print(f"\n   Use this for WatsonX Orchestrate and web-based clients")
        debug_print("="*60 + "\n")

        # FastMCP's run() method with transport="sse" for HTTP mode
        mcp.run(transport="sse", host="0.0.0.0", port=port)
    else:
        # Default stdio mode for Claude Desktop
        debug_print(f"\n📡 Starting stdio MCP server (default)")
        debug_print(f"   Transport: Standard Input/Output")
        debug_print(f"\n   Use this for Claude Desktop and stdio-based clients")
        debug_print(f"   For HTTP/SSE mode, run: python3 {__file__} --http")
        debug_print("="*60 + "\n")

        mcp.run()
